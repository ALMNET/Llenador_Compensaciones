import re
import mpmath as mp
from datetime import date, timedelta, datetime
from openpyxl import Workbook, load_workbook
from colorama import Fore
from bs4 import BeautifulSoup
import requests 


##########################################################################################


url_datos = "https://backend.jaha.com.py/compensaciones/tarjetas_diferencias_eventos"

# url_ingresos_ps = "https://backoffice.jaha.com.py/reportes/ingresospsdetallado"

# client = requests.Session()

# html = client.get('https://backoffice.jaha.com.py/reportes/ingresospsdetallado').content
# soup = BeautifulSoup(html, 'html.parser')

# print(soup.find(name="csrf-token").text)




##########################################################################################
################################### OBTENCION DE DATOS ###################################
##########################################################################################

result = requests.get(url_datos)
content = result.text

fecha_referencia = date.today().strftime('%Y-%m-%d')

# Posicion ultima compensacion segun fecha actual (referencia)
posicion_inicio = content.find(fecha_referencia)
datos_ultima_compensacion = content[posicion_inicio:].replace("</pre>", "")

# Iniciamos un indice para barrido de fechas
i = 1

# Bucle que obtiene fecha y datos de la compensacion mas reciente, 
# barriendo desde hoy a dias anteriores
while datos_ultima_compensacion == ">": # Se compara con valor de datos_ultima_compensacion cuando no hubo compensaciones en la fecha de hoy
    fecha_compensacion = date.today() - timedelta(days = i)
    fecha_compensacion = fecha_compensacion.strftime('%Y-%m-%d')
    posicion_inicio = content.find(fecha_compensacion)
    datos_ultima_compensacion = content[posicion_inicio:].replace("</pre>", "")
    i = i + 1

print(datos_ultima_compensacion)

##########################################################################################
###################################### MODULO EXCEL ######################################
##########################################################################################

excelName = "Planilla de Tiempos de Compensaciones de PY.xlsx"
hojaName = "Datos Backend"
excelObject = load_workbook(excelName)
hojaObject = excelObject[hojaName]


fecha_index = None  # Valor de referencia para fecha

# Ciclo para obtener el indice de la celda donde esta la ultima compensacion
index_tabla = 7 # Un valor de referencia para reducir la cantidad de iteraciones
while fecha_index != fecha_compensacion:
    index_tabla = index_tabla + 1
    fecha_index = hojaObject['A{index_tabla}'.format(index_tabla = index_tabla)].value.strftime('%Y-%m-%d')

# respaldo_compensacion = datos_ultima_compensacion

# Sustituye espacios multiples por puntos y comas
datos_ultima_compensacion = re.sub(r"\s+", ";", datos_ultima_compensacion)

print(datos_ultima_compensacion)

def extractor_datos(texto):
    textoExtraido = texto[:texto.find(";")]
    texto = texto[texto.find(";") + 1:]

    # hojaObject['{fila}{columna}'.format(fila = fila, columna = columna)] = textoExtraido
 
    return textoExtraido, texto

def rellena_celda(texto, fila, columna):
    # hojaObject['{fila}{columna}'.format(fila = fila, columna = columna)] = texto
    #number_format = "0.00"
    number_format = "#,##0"

    hojaObject.cell(row = columna, column = fila).value = texto
    hojaObject.cell(row = columna, column = fila).number_format = number_format

def rellena_celda_flotante(texto, fila, columna):
    # hojaObject['{fila}{columna}'.format(fila = fila, columna = columna)] = texto
    number_format = "0.00"
    # number_format = "#,##0"

    hojaObject.cell(row = columna, column = fila).value = texto
    hojaObject.cell(row = columna, column = fila).number_format = number_format

fila_index = 1 #65 # 'A'
columna_index = index_tabla

Fecha_base, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Tarjetas_Excluidas_Acumuladas, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Tarjetas_en_Espera_Acumuladas, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Tarjetas_Compensadas, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Compensados, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Pago_TDP_a_EPAS_GS, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Pago_TDP_a_EPAS_USD, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Excluidos_Cantidad, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Excluidos_Porcent, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_En_Espera_Cantidad, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_En_Espera_Porcent, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Listo_Cantidad, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Listo_Porcent, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_en_Error_Cantidad, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_en_Error_Porcent, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Perdidos_Cantidad, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Perdidos_Porcent, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Totales_Cantidad, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)
Eventos_Totales_Porcent, datos_ultima_compensacion = extractor_datos(datos_ultima_compensacion)

# rellena_celda(Fecha_base, chr(fila_index + 0), columna_index)
rellena_celda(int(Tarjetas_Excluidas_Acumuladas), (fila_index + 1), columna_index)
rellena_celda(int(Tarjetas_en_Espera_Acumuladas), (fila_index + 2), columna_index)
rellena_celda(int(Tarjetas_Compensadas), (fila_index + 3), columna_index)
rellena_celda(int(Eventos_Compensados), (fila_index + 4), columna_index)
rellena_celda_flotante(float(Pago_TDP_a_EPAS_GS.replace(",", ".")), (fila_index + 5), columna_index)
rellena_celda((Pago_TDP_a_EPAS_USD), (fila_index + 6), columna_index)
rellena_celda(int(Eventos_Excluidos_Cantidad), (fila_index + 7), columna_index)
rellena_celda_flotante(float(Eventos_Excluidos_Porcent.replace(",", ".")), (fila_index + 8), columna_index)
rellena_celda(int(Eventos_En_Espera_Cantidad), (fila_index + 9), columna_index)
rellena_celda_flotante(float(Eventos_En_Espera_Porcent.replace(",", ".")), (fila_index + 10), columna_index)
rellena_celda(int(Eventos_Listo_Cantidad), (fila_index + 11), columna_index)
rellena_celda_flotante(float(Eventos_Listo_Porcent.replace(",", ".")), (fila_index + 12), columna_index)
rellena_celda(int(Eventos_en_Error_Cantidad), (fila_index + 13), columna_index)
rellena_celda_flotante(float(Eventos_en_Error_Porcent.replace(",", ".")), (fila_index + 14), columna_index)
rellena_celda(int(Eventos_Perdidos_Cantidad), (fila_index + 15), columna_index)
rellena_celda_flotante(float(Eventos_Perdidos_Porcent.replace(",", ".")), (fila_index + 16), columna_index)
rellena_celda(int(Eventos_Totales_Cantidad), (fila_index + 17), columna_index)
rellena_celda_flotante(float(Eventos_Totales_Porcent.replace(",", ".")), (fila_index + 18), columna_index)

# noviembre, diciembre y enero
# multas labradas, multas cobradas, credito vendido

print(Tarjetas_Excluidas_Acumuladas)
print(Tarjetas_en_Espera_Acumuladas)
print(Tarjetas_Compensadas)
# print(Eventos_Compensados)
# print(Pago_TDP_a_EPAS_G)

excelObject.save(excelName)

#print("\n" + result2.text)

